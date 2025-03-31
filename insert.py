import speech_recognition as sr
import openpyxl
import spacy
from plyer import notification

# Initialize spaCy NER model
nlp = spacy.load("en_core_web_sm")

# Initialize the Recognizer
recognizer = sr.Recognizer()

# Create a new Excel workbook or load an existing one
excel_file_path = 'speech_to_text_output.xlsx'  # Replace with your desired Excel file path

try:
    workbook = openpyxl.load_workbook(excel_file_path)
except FileNotFoundError:
    workbook = openpyxl.Workbook()

# Select the active sheet (create one if not already present)
sheet = workbook.active
if sheet.title != 'SpeechTranscription':
    sheet.title = 'SpeechTranscription'

# Function to transcribe speech and extract first name, last name, and entities using NER
def transcribe_speech_to_excel():
    try:
        with sr.Microphone() as source:
            print("Listening for a sentence...")
            audio = recognizer.listen(source)

        # Use Google Web Speech API for transcription
        sentence = recognizer.recognize_google(audio)
        print(f"Transcribed Sentence: {sentence}")

        # Extract entities using spaCy NER
        doc = nlp(sentence)

        # Extract first name
        first_name = ""
        last_name = ""
        for ent in doc.ents:
            if ent.label_ == "PERSON":
                name_parts = ent.text.split()
                if len(name_parts) == 2:
                    first_name, last_name = name_parts
                    break

        print(f"Extracted First Name: {first_name}")
        print(f"Extracted Last Name: {last_name}")

        # Extract location
        location = ""
        for ent in doc.ents:
            if ent.label_ == "GPE":  # GPE refers to geopolitical entity (countries, cities, states)
                location = ent.text
                break

        print(f"Extracted Location: {location}")

        # Extract CGPA
        cgpa = ""
        for word in sentence.split():
            if word.isdigit() or (word.count('.') == 1 and word.replace('.', '').isdigit()):
                cgpa = word
                break

        print(f"Extracted CGPA: {cgpa}")

        # Ask user if they are placed
        notification.notify(
            title="Prompt",
            message="Are you placed? Speak 'Yes' or 'No'.",
            timeout=5
        )

        with sr.Microphone() as source:
            print("Are you placed?")
            audio = recognizer.listen(source)

        placement_status = recognizer.recognize_google(audio).lower()
        print(f"Placement Status: {placement_status}")

        if 'yes' in placement_status:
            placement_status = 'yes'
        elif 'no' in placement_status:
            placement_status = 'no'
        else:
            print("Invalid response. Please respond with 'Yes' or 'No'.")
            return

        # Find the next empty row
        next_row = 1
        while sheet.cell(row=next_row, column=1).value is not None:
            next_row += 1

        # Insert the extracted information into the found row
        if first_name:
            sheet.cell(row=next_row, column=1).value = first_name
        if last_name:
            sheet.cell(row=next_row, column=2).value = last_name
        if location:
            sheet.cell(row=next_row, column=3).value = location
        if cgpa:
            sheet.cell(row=next_row, column=4).value = cgpa
        if placement_status:
            sheet.cell(row=next_row, column=5).value = placement_status

        # Save the Excel file
        workbook.save(excel_file_path)

        # Notify user about the entry
        notification_title = "New Entry Added"
        notification_message = f"Name: {first_name} {last_name}, Location: {location}, CGPA: {cgpa}, Placement Status: {placement_status}"
        notification.notify(
            title=notification_title,
            message=notification_message,
            timeout=5  # Notification will disappear after 5 seconds
        )

    except KeyboardInterrupt:
        print('Terminating the program...')
        exit(0)
    except sr.UnknownValueError:
        print("No speech detected or could not be transcribed.")

if __name__ == "__main__":
    print("SpeechDBQuery")
    print("Press Ctrl+C to stop the program.")
    while True:
        try:
            transcribe_speech_to_excel()
        except Exception as e:
            print(f"An error occurred: {e}")