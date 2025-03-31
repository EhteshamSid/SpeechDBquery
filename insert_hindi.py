import speech_recognition as sr
import openpyxl
import spacy
from plyer import notification
from googletrans import Translator

# Initialize spaCy NER model
nlp = spacy.load("en_core_web_sm")

# Initialize the Recognizer
recognizer = sr.Recognizer()

# Create a new Excel workbook or load an existing one
excel_file_path = 'speech_to_text_output.xlsx'  # Replace with your desired Excel file path

try:
    workbook = openpyxl.load_workbook(excel_file_path)
except FileNotFoundError:
    workbook = openpyxl.Workbook()

# Select the active sheet (create one if not already present)
sheet = workbook.active
if sheet.title != 'SpeechTranscription':
    sheet.title = 'SpeechTranscription'

# Initialize translator
translator = Translator()

# Function to transcribe speech, extract information, translate to Hindi, and write to Excel
def transcribe_speech_to_excel():
    try:
        with sr.Microphone() as source:
            print("Listening for a sentence...")
            audio = recognizer.listen(source)

        # Use Google Web Speech API for transcription
        sentence = recognizer.recognize_google(audio)
        print(f"Transcribed Sentence: {sentence}") 

        # Extract entities using spaCy NER
        doc = nlp(sentence)

        # Extract first name
        first_name = ""
        last_name = ""
        for ent in doc.ents:
            if ent.label_ == "PERSON":
                name_parts = ent.text.split()
                if len(name_parts) == 2:
                    first_name, last_name = name_parts
                    break

        print(f"Extracted First Name: {first_name}")
        print(f"Extracted Last Name: {last_name}")

        # Extract location
        location = ""
        for ent in doc.ents:
            if ent.label_ == "GPE":  # GPE refers to geopolitical entity (countries, cities, states)
                location = ent.text
                break

        print(f"Extracted Location: {location}")

        # Extract CGPA
        cgpa = ""
        for word in sentence.split():
            if word.isdigit() or (word.count('.') == 1 and word.replace('.', '').isdigit()):
                cgpa = word
                break

        print(f"Extracted CGPA: {cgpa}")

        # Ask user for placement status
        while True:
            print("Are you placed? Speak 'Yes' or 'No'.")
            with sr.Microphone() as source:
                audio = recognizer.listen(source)

            try:
                placement_status = recognizer.recognize_google(audio).lower()
                print(f"Placement Status: {placement_status}")

                if 'yes' in placement_status:
                    placement_status = 'yes'
                    break
                elif 'no' in placement_status:
                    placement_status = 'no'
                    break
                else:
                    print("Invalid response. Please respond with 'Yes' or 'No'.")
            except sr.UnknownValueError:
                print("No speech detected or could not be transcribed.")

        # Translate extracted information to Hindi
        translated_first_name = translator.translate(first_name, src='en', dest='hi').text
        translated_last_name = translator.translate(last_name, src='en', dest='hi').text
        translated_location = translator.translate(location, src='en', dest='hi').text
        translated_placement_status = translator.translate(placement_status, src='en', dest='hi').text

        # Find the next empty row
        next_row = 1
        while sheet.cell(row=next_row, column=1).value is not None:
            next_row += 1

        # Insert the extracted information into the found row
        sheet.cell(row=next_row, column=1).value = translated_first_name
        sheet.cell(row=next_row, column=2).value = translated_last_name
        sheet.cell(row=next_row, column=3).value = translated_location
        sheet.cell(row=next_row, column=4).value = cgpa
        sheet.cell(row=next_row, column=5).value = translated_placement_status

        # Save the Excel file
        workbook.save(excel_file_path)
        
        # Notify user about the entry
        notification_title = "New Entry Added"
        notification_message = f"Name: {translated_first_name} {translated_last_name}, Location: {translated_location}, CGPA: {cgpa}, Placement Status: {translated_placement_status}"
        notification.notify(
            title=notification_title,
            message=notification_message,
            timeout=5  # Notification will disappear after 5 seconds
        )

    except KeyboardInterrupt:
        print('Terminating the program...')
        exit(0)

if __name__ == "__main__":
    print("SpeechDBQuery")
    print("Press Ctrl+C to stop the program.")
    while True:
        try:
            transcribe_speech_to_excel()
        except Exception as e:
            print(f"An error occurred: {e}")