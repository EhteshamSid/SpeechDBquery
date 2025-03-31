import openpyxl
import speech_recognition as sr

# Load the existing Excel workbook
excel_file_path = 'speech_to_text_output.xlsx'  # Replace with your file path
workbook = openpyxl.load_workbook(excel_file_path)

# Select the desired sheet
sheet = workbook['SpeechTranscription']

# Initialize the speech recognizer
recognizer = sr.Recognizer()

# Function to transcribe speech and delete matching entries
def transcribe_and_delete_entries():
    try:
        with sr.Microphone() as source:
            print("Listening for text to delete...")
            audio = recognizer.listen(source)

        # Use Google Web Speech API for transcription
        spoken_text = recognizer.recognize_google(audio)
        print(f"Spoken Text: {spoken_text}")

        # Iterate through rows to find matching entries and delete them
        rows_to_delete = []
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=3):
            first_name_en = row[0].value
            last_name_en = row[1].value
            if first_name_en and last_name_en:
                full_name_en = f"{first_name_en} {last_name_en}".lower()
                if spoken_text.lower() == full_name_en:
                    rows_to_delete.append(row[0].row)

        # Delete matching rows
        for row_index in sorted(rows_to_delete, reverse=True):
            sheet.delete_rows(row_index)

        # Save the workbook
        workbook.save(excel_file_path)

        print(f"Deleted {len(rows_to_delete)} matching entries.")

    except KeyboardInterrupt:
        print('Terminating the program...')
        exit(0)

    except sr.UnknownValueError:
        print("No speech detected or could not be transcribed.")

if __name__ == "__main__":
    print("SpeechDBQuery - Delete Entries in English")
    transcribe_and_delete_entries()