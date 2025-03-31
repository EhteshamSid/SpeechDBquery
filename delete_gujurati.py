import openpyxl
import speech_recognition as sr
from googletrans import Translator

# Load the existing Excel workbook
excel_file_path = 'speech_to_text_output.xlsx'  # Replace with your file path
workbook = openpyxl.load_workbook(excel_file_path)

# Select the desired sheet
sheet = workbook['SpeechTranscription']

# Initialize the speech recognizer
recognizer = sr.Recognizer()

# Initialize translator
translator = Translator()

# Function to transcribe speech and delete matching entries
def transcribe_and_delete_entries():
    try:
        with sr.Microphone() as source:
            print("Listening for text to delete...")
            audio = recognizer.listen(source)

        # Use Google Web Speech API for transcription
        spoken_text = recognizer.recognize_google(audio, language="gu-IN")
        print(f"Spoken Text: {spoken_text}")

        # Translate spoken text to English
        translated_text = translator.translate(spoken_text, src='gu', dest='en').text
        print(f"Translated Text: {translated_text}")

        # Iterate through rows to find matching entries and delete them
        rows_to_delete = []
        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=1, max_col=3):
            first_name_gu = row[0].value
            last_name_gu = row[1].value
            if first_name_gu and last_name_gu:
                first_name_en = translator.translate(first_name_gu, src='gu', dest='en').text
                last_name_en = translator.translate(last_name_gu, src='gu', dest='en').text
                full_name_en = f"{first_name_en} {last_name_en}".lower()
                if translated_text.lower() == full_name_en:
                    rows_to_delete.append(row[0].row)

        # Delete matching rows
        for row_index in sorted(rows_to_delete, reverse=True):
            sheet.delete_rows(row_index)

        # Save the workbook
        workbook.save(excel_file_path)

        print(f"Deleted {len(rows_to_delete)} matching entries.")

    except KeyboardInterrupt:
        print('Terminating the program...')
        exit(0)

    except sr.UnknownValueError:
        print("No speech detected or could not be transcribed.")

if __name__ == "__main__":
    print("SpeechDBQuery - Delete Entries in Gujarati")
    transcribe_and_delete_entries()